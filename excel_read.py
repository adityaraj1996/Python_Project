import openpyxl
import datetime

path = "C:\\Users\\ARAJ17\\Desktop\\Book1.xlsx"

path2 = "C:\\Users\\ARAJ17\\Documents\\ECap production Calendar 2020.xlsx"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
max_col = sheet_obj.max_column
max_row = sheet_obj.max_row
# fetching today's date
t = datetime.date.today()
# converting date into string
t2 = t.strftime('%m-%d-%Y')
# slicing to extract the date
t3 = t2[3:5:1]
# converting sliced string to int
row_target = int(t3) + 1

for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row=row_target, column=i)
    # val = cell_obj.value, end=' '
    print(cell_obj.value, end=' ')
# dat = str(sheet_obj.cell(row = row_target ,column= 1))
# seg = (sheet_obj.cell(row = row_target ,column= 2))
# act = (sheet_obj.cell(row = row_target ,column= 3))
# map(lambda x,y,z : str(x) + " " + str(y) + " " + str(z),dat*len(seg),seg,act)


# print(datetime.date.today())

# for i in range(1, max_row + 1):
#     for j in range(1, max_col + 1):
#         cell_obj = sheet_obj.cell(row = i, column = j)
#         print(cell_obj.value, '\n')
# c_obj = sheet_obj.cell(row=2, column=1)

# Print value of cell object
# using the value attribute
# print(c_obj.value)
